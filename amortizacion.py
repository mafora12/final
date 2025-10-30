#!/usr/bin/env python3


from datetime import date, timedelta
import math
import csv
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------------
# Funciones de conversión de tasas
# ----------------------------------

def nominal_to_effective_annual(nominal_rate, comp_per_year):
    return (1 + nominal_rate / comp_per_year) ** comp_per_year - 1

def effective_annual_to_period_rate(eff_annual, payments_per_year):
    return (1 + eff_annual) ** (1 / payments_per_year) - 1

def anticipada_to_vencida(rate_anticipada):
    return rate_anticipada / (1 - rate_anticipada)

def parse_rate(rate_value, rate_type, rate_kind, nominal_comp_per_year, payments_per_year):
    if rate_type == "nominal":
        ear = nominal_to_effective_annual(rate_value, nominal_comp_per_year)
        i = effective_annual_to_period_rate(ear, payments_per_year)
    else:
        i = effective_annual_to_period_rate(rate_value, payments_per_year)

    if rate_kind == "anticipada":
        i = anticipada_to_vencida(i)
    return i

# ----------------------------------
# Cálculo de tabla método francés
# ----------------------------------

def cuota_frances(principal, tasa, n_periodos):
    if tasa == 0:
        return principal / n_periodos
    return principal * (tasa * (1 + tasa) ** n_periodos) / ((1 + tasa) ** n_periodos - 1)

def generar_tabla(principal, tasa, n_periodos, frecuencia, abonos, reducir):
    saldo = principal
    cuota = cuota_frances(principal, tasa, n_periodos)
    tabla = []
    fecha = date.today()

    for periodo in range(1, n_periodos + 1):
        interes = saldo * tasa
        abono_capital = cuota - interes
        abono_extra = abonos.get(periodo, 0)
        saldo -= (abono_capital + abono_extra)

        if saldo < 0:
            saldo = 0

        tabla.append({
            "Periodo": periodo,
            "Fecha": fecha.strftime("%d/%m/%Y"),
            "Cuota ($)": round(cuota, 2),
            "Interés ($)": round(interes, 2),
            "Abono a Capital ($)": round(abono_capital, 2),
            "Abono Extra ($)": round(abono_extra, 2),
            "Saldo Restante ($)": round(saldo, 2)
        })

        if abono_extra > 0 and reducir == "plazo":
            n_restante = n_periodos - periodo
            cuota = cuota_frances(saldo, tasa, n_restante)

        fecha += timedelta(days=30)

        if saldo <= 1e-6:
            break

    return tabla

# ----------------------------------
# Exportar CSV y XLSX
# ----------------------------------

def exportar_archivos(tabla):
    csv_name = "tabla_amortizacion.csv"
    with open(csv_name, "w", newline="", encoding="utf-8-sig") as f:
        escritor = csv.DictWriter(f, fieldnames=tabla[0].keys())
        escritor.writeheader()
        escritor.writerows(tabla)
    print(f"\n✅ Archivo CSV generado: {csv_name}")

    df = pd.DataFrame(tabla)
    xlsx_name = "tabla_amortizacion.xlsx"

    plt.figure(figsize=(8, 4))
    plt.bar(df["Periodo"], df["Cuota ($)"], color="#009FE3", label="Cuota total")
    plt.bar(df["Periodo"], df["Interés ($)"], color="#F4B400", label="Interés")
    plt.bar(df["Periodo"], df["Abono a Capital ($)"], color="#34A853", label="Abono capital")
    plt.title("Distribución de Pagos - Simulación de Crédito")
    plt.xlabel("Periodo")
    plt.ylabel("Valor ($)")
    plt.legend()
    plt.tight_layout()
    grafico = "grafico_pagos.png"
    plt.savefig(grafico, dpi=150)
    plt.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla de Amortización"

    ws["A1"] = "Tabla de Amortización - Simulación de Crédito"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    img = Image(grafico)
    img.width = 600
    img.height = 300
    ws.add_image(img, f"A{len(df) + 4}")

    wb.save(xlsx_name)
    print(f"✅ Archivo XLSX generado: {xlsx_name}")

# ----------------------------------
# Programa principal con validaciones
# ----------------------------------

def pedir_float(mensaje):
    while True:
        try:
            valor = input(mensaje).replace(",", ".")
            return float(valor)
        except ValueError:
            print("❌ Entrada inválida. Usa solo números y puntos decimales (ej. 3.5). Intenta de nuevo.")

def pedir_int(mensaje):
    while True:
        try:
            return int(input(mensaje))
        except ValueError:
            print("❌ Entrada inválida. Debe ser un número entero. Intenta de nuevo.")

def main():
    print("=== SIMULADOR DE CRÉDITO - TABLA DE AMORTIZACIÓN ===\n")

    principal = pedir_float("Monto del crédito ($): ")
    tasa_valor = pedir_float("Tasa de interés (con coma o punto) (%): ") / 100
    tipo_tasa = input("Tipo de tasa (nominal/efectiva): ").strip().lower()
    clase_tasa = input("Clase de tasa (vencida/anticipada): ").strip().lower()
    capitalizacion = pedir_int("Capitalizaciones por año (ej. 12 para mensual): ")
    frecuencia = pedir_int("Pagos por año (12 mensual, 4 trimestral, etc.): ")
    plazo = pedir_int("Plazo total en meses: ")

    tasa_periodo = parse_rate(tasa_valor, tipo_tasa, clase_tasa, capitalizacion, frecuencia)
    print(f"\nTasa por periodo: {tasa_periodo*100:.4f}%")

    abonos = {}
    print("\n¿Deseas ingresar abonos extra? (s/n)")
    if input().strip().lower() == "s":
        while True:
            p = pedir_int("Periodo del abono: ")
            monto = pedir_float("Monto del abono ($): ")
            abonos[p] = monto
            if input("¿Otro abono? (s/n): ").strip().lower() != "s":
                break

    # ✅ Solo preguntar “reducir plazo o cuota” si hay abonos
    reducir = "cuota"
    if abonos:
        reducir = input("\nTras abono extra, ¿reducir 'plazo' o 'cuota'?: ").strip().lower()

    tabla = generar_tabla(principal, tasa_periodo, plazo, frecuencia, abonos, reducir)

    print("\n=== TABLA DE AMORTIZACIÓN ===")
    for fila in tabla:
        print(fila)

    exportar_archivos(tabla)

    saldo_final = tabla[-1]["Saldo Restante ($)"]
    print(f"\nSaldo final: ${saldo_final:.2f}")

if __name__ == "__main__":
    main()
